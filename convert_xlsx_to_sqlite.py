import os
import sqlite3
import openpyxl

def main():
    xlsx_path = "movies.xlsx"
    db_path = "sqlite3.db"
    
    if not os.path.exists(xlsx_path):
        print(f"Error: {xlsx_path} not found.")
        return
        
    print(f"Reading data from {xlsx_path}...")
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        # Use 'All Movies' sheet if it exists, otherwise active sheet
        sheet_name = "All Movies" if "All Movies" in wb.sheetnames else wb.sheetnames[0]
        sheet = wb[sheet_name]
        
        # Read all rows
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            print("Error: Excel file is empty.")
            return
            
        headers = rows[0]
        data_rows = rows[1:]
        print(f"Loaded {len(data_rows)} rows from sheet '{sheet_name}'.")
        
        # Connect to SQLite
        print(f"Connecting to SQLite database at {db_path}...")
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        # Drop table if exists
        cursor.execute("DROP TABLE IF EXISTS movies")
        
        # Create Table
        cursor.execute("""
            CREATE TABLE movies (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                poster_filename TEXT,
                title TEXT,
                categories TEXT,
                score REAL,
                release_date TEXT,
                duration TEXT,
                regions TEXT,
                detail_link TEXT
            )
        """)
        
        # Insert Data
        insert_query = """
            INSERT INTO movies (
                poster_filename, title, categories, score, release_date, duration, regions, detail_link
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """
        
        count = 0
        for row in data_rows:
            # Skip empty rows
            if not any(row):
                continue
                
            # Pad row if columns are missing
            row_padded = list(row) + [None] * (8 - len(row))
            # Extract fields
            poster_filename = row_padded[0]
            title = row_padded[1]
            categories = row_padded[2]
            score = row_padded[3]
            release_date = row_padded[4]
            duration = row_padded[5]
            regions = row_padded[6]
            detail_link = row_padded[7]
            
            # Cast score to float safely
            try:
                score = float(score) if score is not None else 0.0
            except ValueError:
                score = 0.0
                
            cursor.execute(insert_query, (
                poster_filename, title, categories, score, release_date, duration, regions, detail_link
            ))
            count += 1
            
        conn.commit()
        conn.close()
        
        print(f"Successfully converted {count} movies and saved to database: {db_path}")
        
    except Exception as e:
        print(f"Error during conversion: {e}")

if __name__ == "__main__":
    main()
